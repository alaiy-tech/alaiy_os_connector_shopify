"""
Combines the per-slice CSVs written by scan_supplier_totals.py into one
Excel workbook with three sheets, instead of leaving several flat CSVs
to stitch together by hand:

  - Summary: one row per supplier -- local vs Shopify product counts,
    Shopify orders in the window, how many matched locally.
  - Product Mismatches: only suppliers where local_products !=
    shopify_products, for a focused review list.
  - Missing Orders: one row per individual missing order (supplier +
    Shopify order name), flattened out of each supplier's semicolon-
    joined list so it's filterable/sortable in Excel.

Reads whatever supplier_totals*.csv files already exist in the site's
private/files (both the sliced and the non-sliced single-file shape),
so it works whether scan_supplier_totals.py was run in one process or
split across parallel tmux sessions. Makes no other writes; the output
Excel is saved to both private/files and public/files (downloadable).

  bench --site <site> execute \
      alaiy_os_connector_shopify.shopify.combine_supplier_totals.run
"""

import csv
import glob
import shutil

import frappe


def run():
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    site_files_dir = frappe.utils.get_site_path("private/files")
    csv_paths = sorted(glob.glob(f"{site_files_dir}/supplier_totals*.csv"))
    if not csv_paths:
        print("No supplier_totals*.csv files found -- run scan_supplier_totals.py first.")
        return

    print(f"Combining {len(csv_paths)} file(s): {[p.split('/')[-1] for p in csv_paths]}")

    rows = []
    for path in csv_paths:
        with open(path, newline="", encoding="utf-8") as f:
            rows.extend(csv.DictReader(f))

    wb = openpyxl.Workbook()

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    header_font = Font(bold=True)

    # Summary
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.append(["Supplier", "Local Products", "Shopify Products", "Product Mismatch",
                "Shopify Orders In Window", "Local Orders Matched", "Missing Orders Count"])
    for row in sorted(rows, key=lambda r: r["supplier"]):
        r_idx = ws1.max_row + 1
        ws1.append([
            row["supplier"], row["local_products"], row["shopify_products"], row["product_mismatch"],
            row["shopify_orders_in_window"], row["local_orders_matched"], row["missing_orders_count"],
        ])
        if row["product_mismatch"] == "True" or int(row["missing_orders_count"] or 0) > 0:
            for col in range(1, 8):
                ws1.cell(row=r_idx, column=col).fill = red_fill

    # Product Mismatches
    ws2 = wb.create_sheet("Product Mismatches")
    ws2.append(["Supplier", "Local Products", "Shopify Products", "Difference"])
    for row in rows:
        if row["product_mismatch"] != "True":
            continue
        local = int(row["local_products"] or 0)
        shopify = row["shopify_products"]
        diff = (local - int(shopify)) if shopify not in ("unverified", "") else "N/A (unverified)"
        ws2.append([row["supplier"], local, shopify, diff])

    # Missing Orders -- flatten the semicolon-joined list into one row per order
    ws3 = wb.create_sheet("Missing Orders")
    ws3.append(["Supplier", "Shopify Order Name"])
    for row in rows:
        for order_name in (row.get("missing_orders") or "").split(";"):
            order_name = order_name.strip()
            if order_name:
                ws3.append([row["supplier"], order_name])

    for ws in (ws1, ws2, ws3):
        for cell in ws[1]:
            cell.font = header_font
        for col_cells in ws.columns:
            length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(length + 2, 50)

    filename = "supplier_totals_combined.xlsx"
    private_path = frappe.utils.get_site_path(f"private/files/{filename}")
    wb.save(private_path)
    public_path = frappe.utils.get_site_path(f"public/files/{filename}")
    shutil.copy(private_path, public_path)

    print(f"\nWrote {len(rows)} supplier row(s), "
          f"{sum(1 for r in rows if r['product_mismatch'] == 'True')} product mismatch(es), "
          f"{sum(len((r.get('missing_orders') or '').split(';')) if r.get('missing_orders') else 0 for r in rows)} missing order row(s) "
          f"to {private_path}")
    print(f"Downloadable at /files/{filename}")

    return {"suppliers": len(rows), "output": public_path}
