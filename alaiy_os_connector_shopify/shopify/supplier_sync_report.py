"""
Supplier Sync Report: one row per supplier login, showing Product Sync
and Order Sync status -- for sharing with AJ, updated as issues get
resolved.

Product Sync: FAILED if local product count != Shopify's real product
count (from real inventory data via
scan_live_location_product_counts.py's output -- NOT the broken
productsCount(query: "inventory_location_id:...") filter, which
returned the same nonsense round number for every supplier).

Order Sync: FAILED if Shopify has an order in the date window with no
matching local Sales Order at all (from scan_supplier_totals.py's
output).

Reads both scripts' already-written CSVs rather than re-scanning --
run scan_live_location_product_counts.py and scan_supplier_totals.py
(each possibly sliced across parallel sessions) first, then this.

  bench --site <site> execute \
      alaiy_os_connector_shopify.shopify.supplier_sync_report.run
"""

import csv
import glob
import shutil

import frappe


def run():
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    site_files_dir = frappe.utils.get_site_path("private/files")

    location_csvs = sorted(glob.glob(f"{site_files_dir}/live_location_product_counts*.csv"))
    order_csvs = sorted(glob.glob(f"{site_files_dir}/supplier_totals*.csv"))

    if not location_csvs:
        print("No live_location_product_counts*.csv found -- run scan_live_location_product_counts.py first.")
        return
    if not order_csvs:
        print("No supplier_totals*.csv found -- run scan_supplier_totals.py first.")
        return

    print(f"Reading {len(location_csvs)} product-count file(s), {len(order_csvs)} order file(s)")

    # Real Shopify product counts, per supplier (via location -> linked_supplier)
    gid_to_supplier = {
        r.sh_location_gid: r.linked_supplier
        for r in frappe.get_all(
            "Shopify Location", filters={"linked_supplier": ["!=", ""]},
            fields=["sh_location_gid", "linked_supplier"],
        )
        if r.sh_location_gid
    }
    shopify_products_by_supplier = {}
    for path in location_csvs:
        with open(path, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                supplier = gid_to_supplier.get(row["location_gid"])
                if supplier:
                    shopify_products_by_supplier[supplier] = (
                        shopify_products_by_supplier.get(supplier, 0) + int(row["product_count"] or 0)
                    )

    # Local product counts + order data, per supplier
    order_rows = {}
    for path in order_csvs:
        with open(path, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                order_rows[row["supplier"]] = row

    all_suppliers = sorted(set(order_rows) | set(shopify_products_by_supplier))

    report_rows = []
    for supplier in all_suppliers:
        order_row = order_rows.get(supplier, {})
        local_products = int(order_row.get("local_products") or 0)
        shopify_products = shopify_products_by_supplier.get(supplier, 0)
        product_sync_failed = local_products != shopify_products

        shopify_orders = int(order_row.get("shopify_orders_in_window") or 0)
        local_orders = int(order_row.get("local_orders_matched") or 0)
        missing_orders_count = int(order_row.get("missing_orders_count") or 0)
        order_sync_failed = missing_orders_count > 0
        missing_orders = [o.strip() for o in (order_row.get("missing_orders") or "").split(";") if o.strip()]

        report_rows.append({
            "supplier": supplier,
            "local_products": local_products,
            "shopify_products": shopify_products,
            "product_sync_status": "FAILED" if product_sync_failed else "OK",
            "shopify_orders": shopify_orders,
            "local_orders": local_orders,
            "missing_orders_count": missing_orders_count,
            "order_sync_status": "FAILED" if order_sync_failed else "OK",
            "missing_orders": missing_orders,
        })

    wb = openpyxl.Workbook()
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    header_font = Font(bold=True)

    ws1 = wb.active
    ws1.title = "Supplier Sync Report"
    ws1.append(["Supplier", "Local Products", "Shopify Products", "Product Sync",
                "Shopify Orders", "Local Orders Matched", "Missing Orders", "Order Sync"])
    for row in report_rows:
        r_idx = ws1.max_row + 1
        ws1.append([
            row["supplier"], row["local_products"], row["shopify_products"], row["product_sync_status"],
            row["shopify_orders"], row["local_orders"], row["missing_orders_count"], row["order_sync_status"],
        ])
        product_col_fill = red_fill if row["product_sync_status"] == "FAILED" else green_fill
        order_col_fill = red_fill if row["order_sync_status"] == "FAILED" else green_fill
        ws1.cell(row=r_idx, column=4).fill = product_col_fill
        ws1.cell(row=r_idx, column=8).fill = order_col_fill

    ws2 = wb.create_sheet("Failed Product Sync")
    ws2.append(["Supplier", "Local Products", "Shopify Products", "Difference"])
    for row in report_rows:
        if row["product_sync_status"] == "FAILED":
            ws2.append([row["supplier"], row["local_products"], row["shopify_products"],
                        row["local_products"] - row["shopify_products"]])

    ws3 = wb.create_sheet("Failed Order Sync")
    ws3.append(["Supplier", "Missing Shopify Order"])
    for row in report_rows:
        for order_name in row["missing_orders"]:
            ws3.append([row["supplier"], order_name])

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = header_font
        for col_cells in ws.columns:
            length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(length + 2, 50)

    filename = "supplier_sync_report.xlsx"
    private_path = frappe.utils.get_site_path(f"private/files/{filename}")
    wb.save(private_path)
    public_path = frappe.utils.get_site_path(f"public/files/{filename}")
    shutil.copy(private_path, public_path)

    failed_product = sum(1 for r in report_rows if r["product_sync_status"] == "FAILED")
    failed_order = sum(1 for r in report_rows if r["order_sync_status"] == "FAILED")
    print(f"\n{len(report_rows)} supplier(s) checked.")
    print(f"Product Sync FAILED: {failed_product}")
    print(f"Order Sync FAILED: {failed_order}")
    print(f"\nWrote report to {private_path}")
    print(f"Downloadable at /files/{filename}")

    return {"suppliers": len(report_rows), "product_sync_failed": failed_product, "order_sync_failed": failed_order}
