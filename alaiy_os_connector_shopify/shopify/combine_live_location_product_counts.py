"""
The one combined per-supplier report: products AND orders, local vs
real Shopify data, side by side.

Reads two sets of already-written CSVs (run both scan scripts first,
each possibly sliced across parallel sessions):
  - live_location_product_counts*.csv (from
    scan_live_location_product_counts.py) -- real Shopify product
    counts, from actual inventory data. NOT the broken
    productsCount(query: "inventory_location_id:...") filter, which
    returned the same nonsense round number for every supplier.
  - supplier_totals*.csv (from scan_supplier_totals.py) -- Shopify
    orders since a given date, matched against local Sales Orders,
    and local product counts.

  bench --site <site> execute \
      alaiy_os_connector_shopify.shopify.combine_live_location_product_counts.run
"""

import csv
import glob
import shutil

import frappe


def run():
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    site_files_dir = frappe.utils.get_site_path("private/files")
    location_csvs = sorted(glob.glob(f"{site_files_dir}/live_location_product_counts*.csv"))
    order_csvs = sorted(glob.glob(f"{site_files_dir}/supplier_totals*.csv"))

    if not location_csvs:
        print("No live_location_product_counts*.csv files found -- run scan_live_location_product_counts.py first.")
        return
    if not order_csvs:
        print("No supplier_totals*.csv files found -- run scan_supplier_totals.py first.")
        return

    print(f"Combining {len(location_csvs)} product-count file(s), {len(order_csvs)} order file(s)")

    location_rows = []
    for path in location_csvs:
        with open(path, newline="", encoding="utf-8") as f:
            location_rows.extend(csv.DictReader(f))

    gid_to_supplier = {
        r.sh_location_gid: r.linked_supplier
        for r in frappe.get_all(
            "Shopify Location", filters={"linked_supplier": ["!=", ""]},
            fields=["sh_location_gid", "linked_supplier"],
        )
        if r.sh_location_gid
    }

    supplier_shopify_products = {}
    unmapped_locations = []
    for row in location_rows:
        supplier = gid_to_supplier.get(row["location_gid"])
        count = int(row["product_count"] or 0)
        if supplier:
            supplier_shopify_products[supplier] = supplier_shopify_products.get(supplier, 0) + count
        else:
            unmapped_locations.append(row)

    order_rows = {}
    for path in order_csvs:
        with open(path, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                order_rows[row["supplier"]] = row

    all_suppliers = sorted(set(order_rows) | set(supplier_shopify_products))

    wb = openpyxl.Workbook()
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    header_font = Font(bold=True)

    ws1 = wb.active
    ws1.title = "Summary"
    ws1.append(["Supplier", "Local Products", "Shopify Products (real stock)", "Product Mismatch",
                "Shopify Orders", "Local Orders Matched", "Missing Orders"])
    for supplier in all_suppliers:
        order_row = order_rows.get(supplier, {})
        local_products = int(order_row.get("local_products") or 0)
        shopify_products = supplier_shopify_products.get(supplier, 0)
        product_mismatch = local_products != shopify_products

        shopify_orders = int(order_row.get("shopify_orders_in_window") or 0)
        local_orders = int(order_row.get("local_orders_matched") or 0)
        missing_orders_count = int(order_row.get("missing_orders_count") or 0)

        r_idx = ws1.max_row + 1
        ws1.append([supplier, local_products, shopify_products, product_mismatch,
                    shopify_orders, local_orders, missing_orders_count])
        if product_mismatch:
            for col in (1, 2, 3, 4):
                ws1.cell(row=r_idx, column=col).fill = red_fill
        if missing_orders_count > 0:
            for col in (5, 6, 7):
                ws1.cell(row=r_idx, column=col).fill = red_fill

    ws2 = wb.create_sheet("Missing Orders")
    ws2.append(["Supplier", "Missing Shopify Order"])
    for supplier, order_row in order_rows.items():
        for order_name in (order_row.get("missing_orders") or "").split(";"):
            order_name = order_name.strip()
            if order_name:
                ws2.append([supplier, order_name])

    ws3 = wb.create_sheet("By Location")
    ws3.append(["Location GID", "Location Name", "Product Count", "Linked Supplier"])
    for row in sorted(location_rows, key=lambda r: -int(r["product_count"] or 0)):
        ws3.append([row["location_gid"], row["location_name"], row["product_count"],
                    gid_to_supplier.get(row["location_gid"], "")])

    if unmapped_locations:
        ws4 = wb.create_sheet("Unmapped Locations")
        ws4.append(["Location GID", "Location Name", "Product Count"])
        for row in unmapped_locations:
            ws4.append([row["location_gid"], row["location_name"], row["product_count"]])

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = header_font
        for col_cells in ws.columns:
            length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(length + 2, 50)

    filename = "supplier_products_orders_report.xlsx"
    private_path = frappe.utils.get_site_path(f"private/files/{filename}")
    wb.save(private_path)
    public_path = frappe.utils.get_site_path(f"public/files/{filename}")
    shutil.copy(private_path, public_path)

    product_mismatches = sum(
        1 for s in all_suppliers
        if int(order_rows.get(s, {}).get("local_products") or 0) != supplier_shopify_products.get(s, 0)
    )
    total_missing_orders = sum(int(order_rows.get(s, {}).get("missing_orders_count") or 0) for s in all_suppliers)

    print(f"\n{len(all_suppliers)} supplier(s). {product_mismatches} product count mismatch(es). "
          f"{total_missing_orders} missing order(s) total. {len(unmapped_locations)} unmapped location(s).")
    print(f"Wrote report to {private_path}")
    print(f"Downloadable at /files/{filename}")

    return {"suppliers": len(all_suppliers), "product_mismatches": product_mismatches,
            "missing_orders": total_missing_orders}
